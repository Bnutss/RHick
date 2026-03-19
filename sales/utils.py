from openpyxl.styles import Font, Alignment
from openpyxl.drawing.image import Image
from openpyxl import Workbook
from PIL import Image as PILImage
import os
import requests
import mimetypes
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfbase import pdfmetrics
from reportlab.platypus import SimpleDocTemplate, Table, Spacer, TableStyle, Image as ReportLabImage, Paragraph, \
    HRFlowable
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER

mimetypes.add_type('image/webp', '.webp')

TELEGRAM_BOT_TOKEN = '7775474735:AAFHyJw-YL1e91AIVj-KIrWxg8Ps6GprXhs'

TELEGRAM_CHAT_ID = '-1002411014709'

HIK_RED = colors.HexColor('#E2001A')
HIK_DARK = colors.HexColor('#1A1A1A')
HIK_GRAY = colors.HexColor('#F7F7F7')
HIK_LIGHT = colors.HexColor('#FAFAFA')
HIK_BORDER = colors.HexColor('#E0E0E0')
HIK_TEXT_GRAY = colors.HexColor('#888888')
HIK_FINAL_TEXT = colors.HexColor('#FFFFFF')


def convert_webp_to_png(photo_path):
    png_path = photo_path.replace('.webp', '.png')
    with PILImage.open(photo_path) as img:
        img.save(png_path, 'PNG')
    return png_path


def generate_order_excel(order):
    wb = Workbook()
    ws = wb.active
    ws.title = "Order"

    headers = [
        "Название клиента:", order.client,
        "Использованный НДС:", f"{order.vat}%",
        "Прочие расходы (%):", f"{order.additional_expenses}%"
    ]

    for col in range(0, len(headers), 2):
        cell = ws.cell(row=1 + col // 2, column=1, value=headers[col])
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell = ws.cell(row=1 + col // 2, column=2, value=headers[col + 1])
        cell.alignment = Alignment(horizontal='center', vertical='center')

    table_headers = ["Название товара", "Фото", "Количество", "Цена за единицу", "Общая стоимость"]
    for col_num, header in enumerate(table_headers, 1):
        cell = ws.cell(row=4, column=col_num, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    row_num = 5
    for product in order.products.all():
        ws.cell(row=row_num, column=1, value=product.name).alignment = Alignment(horizontal='center', vertical='center')

        if product.photo:
            photo_path = product.photo.path
            try:
                if photo_path.lower().endswith('.webp'):
                    photo_path = convert_webp_to_png(photo_path)
                img = Image(photo_path)
                img.width = 30
                img.height = 30
                img.anchor = f'B{row_num}'
                ws.add_image(img)
            except Exception as e:
                print(f"Ошибка обработки изображения: {e}")
                ws.cell(row=row_num, column=2, value="Изображение недоступно")

        ws.cell(row=row_num, column=3, value=product.quantity).alignment = Alignment(horizontal='center',
                                                                                     vertical='center')
        ws.cell(row=row_num, column=4, value=product.price).alignment = Alignment(horizontal='center',
                                                                                  vertical='center')
        ws.cell(row=row_num, column=5, value=product.quantity * product.price).alignment = Alignment(
            horizontal='center', vertical='center')
        ws.row_dimensions[row_num].height = 30
        row_num += 1

    total_price = order.get_total_price()
    total_price_with_vat = order.get_total_price_with_vat()
    additional_expenses_amount = order.get_additional_expenses_amount()
    total_sum = total_price_with_vat + additional_expenses_amount

    ws.cell(row=row_num, column=4, value="Итого без НДС:").font = Font(bold=True)
    ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row_num, column=5, value=total_price).font = Font(bold=True)
    ws.cell(row=row_num, column=5).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(row=row_num + 1, column=4, value="Итого с НДС:").font = Font(bold=True)
    ws.cell(row=row_num + 1, column=4).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row_num + 1, column=5, value=total_price_with_vat).font = Font(bold=True)
    ws.cell(row=row_num + 1, column=5).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(row=row_num + 2, column=4, value="Прочие расходы:").font = Font(bold=True)
    ws.cell(row=row_num + 2, column=4).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row_num + 2, column=5, value=additional_expenses_amount).font = Font(bold=True)
    ws.cell(row=row_num + 2, column=5).alignment = Alignment(horizontal='center', vertical='center')

    ws.cell(row=row_num + 3, column=4, value="Общий итог:").font = Font(bold=True)
    ws.cell(row=row_num + 3, column=4).alignment = Alignment(horizontal='center', vertical='center')
    ws.cell(row=row_num + 3, column=5, value=total_sum).font = Font(bold=True)
    ws.cell(row=row_num + 3, column=5).alignment = Alignment(horizontal='center', vertical='center')

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 20

    file_name = f"Заказ-{order.id}.xlsx"
    file_path = os.path.join("/tmp", file_name)
    wb.save(file_path)
    return file_path


def register_fonts():
    font_path = os.path.join(os.path.dirname(__file__), 'static', 'fonts', 'Roboto-Regular.ttf')
    pdfmetrics.registerFont(TTFont('Roboto', font_path))


def convert_image_for_pdf(photo_path):
    if photo_path.lower().endswith('.webp'):
        png_path = photo_path.replace('.webp', '.png')
        with PILImage.open(photo_path) as img:
            img.save(png_path, 'PNG')
        return png_path
    return photo_path


def generate_order_pdf(order):
    register_fonts()

    pdf_file_name = f"Заказ-{order.id}.pdf"
    pdf_file_path = os.path.join("/tmp", pdf_file_name)
    doc = SimpleDocTemplate(
        pdf_file_path,
        pagesize=A4,
        rightMargin=1 * cm,
        leftMargin=1 * cm,
        topMargin=1 * cm,
        bottomMargin=2 * cm
    )
    elements = []

    header_data = [[
        Table(
            [[
                Paragraph(f'<font name="Roboto" size="14" color="#1A1A1A"><b>RHIK</b></font>',
                          ParagraphStyle('', fontName='Roboto')),
                Paragraph(f'<font name="Roboto" size="7" color="#888888">Security &amp; Systems</font>',
                          ParagraphStyle('', fontName='Roboto')),
            ]],
            colWidths=[3 * cm, 6 * cm],
            style=[
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ]
        ),
        Table(
            [[
                Paragraph(f'<font name="Roboto" size="14" color="#E2001A"><b>Заказ #{order.id}</b></font>',
                          ParagraphStyle('', fontName='Roboto', alignment=TA_RIGHT)),
            ]],
            colWidths=[9 * cm],
            style=[
                ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]
        ),
    ]]

    header_table = Table(header_data, colWidths=[9 * cm, 9 * cm])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LINEBELOW', (0, 0), (-1, -1), 2, HIK_RED),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 0.3 * cm))

    if order.is_confirmed:
        status_text = 'Подтвержден'
        status_bg = colors.HexColor('#D4EDDA')
        status_color = colors.HexColor('#155724')
    elif order.is_rejected:
        status_text = 'Отклонен'
        status_bg = colors.HexColor('#F8D7DA')
        status_color = colors.HexColor('#721C24')
    else:
        status_text = 'В ожидании'
        status_bg = colors.HexColor('#FFF3CD')
        status_color = colors.HexColor('#856404')

    info_data = [[
        Table([
            [Paragraph('<font name="Roboto" size="5" color="#888888">КЛИЕНТ</font>',
                       ParagraphStyle('', fontName='Roboto'))],
            [Paragraph(f'<font name="Roboto" size="8" color="#1A1A1A"><b>{order.client}</b></font>',
                       ParagraphStyle('', fontName='Roboto'))],
        ], colWidths=[4.2 * cm], style=[('LEFTPADDING', (0, 0), (-1, -1), 6), ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                                        ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                                        ('LINEBEFORE', (0, 0), (0, -1), 3, HIK_RED),
                                        ('BACKGROUND', (0, 0), (-1, -1), HIK_GRAY)]),
        Table([
            [Paragraph('<font name="Roboto" size="5" color="#888888">НДС</font>',
                       ParagraphStyle('', fontName='Roboto'))],
            [Paragraph(f'<font name="Roboto" size="8" color="#1A1A1A"><b>{order.vat if order.vat else 0}%</b></font>',
                       ParagraphStyle('', fontName='Roboto'))],
        ], colWidths=[4.2 * cm], style=[('LEFTPADDING', (0, 0), (-1, -1), 6), ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                                        ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                                        ('LINEBEFORE', (0, 0), (0, -1), 3, HIK_RED),
                                        ('BACKGROUND', (0, 0), (-1, -1), HIK_GRAY)]),
        Table([
            [Paragraph('<font name="Roboto" size="5" color="#888888">ПРОЧИЕ РАСХОДЫ</font>',
                       ParagraphStyle('', fontName='Roboto'))],
            [Paragraph(
                f'<font name="Roboto" size="8" color="#1A1A1A"><b>{order.additional_expenses if order.additional_expenses else 0}%</b></font>',
                ParagraphStyle('', fontName='Roboto'))],
        ], colWidths=[4.2 * cm], style=[('LEFTPADDING', (0, 0), (-1, -1), 6), ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                                        ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                                        ('LINEBEFORE', (0, 0), (0, -1), 3, HIK_RED),
                                        ('BACKGROUND', (0, 0), (-1, -1), HIK_GRAY)]),
        Table([
            [Paragraph('<font name="Roboto" size="5" color="#888888">СТАТУС</font>',
                       ParagraphStyle('', fontName='Roboto'))],
            [Paragraph(f'<font name="Roboto" size="7" color="#{status_color.hexval()[2:]}"><b>{status_text}</b></font>',
                       ParagraphStyle('', fontName='Roboto'))],
        ], colWidths=[4.2 * cm], style=[('LEFTPADDING', (0, 0), (-1, -1), 6), ('RIGHTPADDING', (0, 0), (-1, -1), 4),
                                        ('TOPPADDING', (0, 0), (-1, -1), 4), ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                                        ('LINEBEFORE', (0, 0), (0, -1), 3, HIK_RED),
                                        ('BACKGROUND', (0, 0), (-1, -1), status_bg)]),
    ]]

    info_table = Table(info_data, colWidths=[4.4 * cm, 4.4 * cm, 4.4 * cm, 4.4 * cm])
    info_table.setStyle(TableStyle([
        ('LEFTPADDING', (0, 0), (-1, -1), 2),
        ('RIGHTPADDING', (0, 0), (-1, -1), 2),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 0.4 * cm))

    elements.append(HRFlowable(width='100%', thickness=1, color=HIK_RED, spaceAfter=6))

    elements.append(Paragraph(
        '<font name="Roboto" size="7" color="#E2001A"><b>СПИСОК ТОВАРОВ</b></font>',
        ParagraphStyle('', fontName='Roboto', spaceBefore=2, spaceAfter=6)
    ))

    product_header = [
        Paragraph('<font name="Roboto" size="6" color="#FFFFFF"><b>НАЗВАНИЕ ТОВАРА</b></font>',
                  ParagraphStyle('', fontName='Roboto', alignment=TA_LEFT)),
        Paragraph('<font name="Roboto" size="6" color="#FFFFFF"><b>ФОТО</b></font>',
                  ParagraphStyle('', fontName='Roboto', alignment=TA_CENTER)),
        Paragraph('<font name="Roboto" size="6" color="#FFFFFF"><b>КОЛ-ВО</b></font>',
                  ParagraphStyle('', fontName='Roboto', alignment=TA_CENTER)),
        Paragraph('<font name="Roboto" size="6" color="#FFFFFF"><b>ЦЕНА ЗА ЕД.</b></font>',
                  ParagraphStyle('', fontName='Roboto', alignment=TA_CENTER)),
        Paragraph('<font name="Roboto" size="6" color="#FFFFFF"><b>ОБЩАЯ СТОИМОСТЬ</b></font>',
                  ParagraphStyle('', fontName='Roboto', alignment=TA_CENTER)),
    ]

    data = [product_header]

    for i, product in enumerate(order.products.all()):
        row_bg = HIK_LIGHT if i % 2 == 0 else colors.white

        photo_cell = Paragraph('<font name="Roboto" size="6" color="#BBBBBB">Нет</font>',
                               ParagraphStyle('', fontName='Roboto', alignment=TA_CENTER))
        if product.photo:
            try:
                photo_path = convert_image_for_pdf(product.photo.path)
                if os.path.exists(photo_path):
                    photo_cell = ReportLabImage(photo_path, width=1 * cm, height=1 * cm)
            except Exception:
                pass

        row = [
            Paragraph(f'<font name="Roboto" size="7" color="#1A1A1A"><b>{product.name}</b></font>',
                      ParagraphStyle('', fontName='Roboto', alignment=TA_LEFT)),
            photo_cell,
            Paragraph(f'<font name="Roboto" size="7" color="#333333">{product.quantity}</font>',
                      ParagraphStyle('', fontName='Roboto', alignment=TA_CENTER)),
            Paragraph(f'<font name="Roboto" size="7" color="#333333">{product.price:.2f}</font>',
                      ParagraphStyle('', fontName='Roboto', alignment=TA_CENTER)),
            Paragraph(
                f'<font name="Roboto" size="7" color="#E2001A"><b>{product.quantity * product.price:.2f}</b></font>',
                ParagraphStyle('', fontName='Roboto', alignment=TA_CENTER)),
        ]
        data.append(row)

    products_table = Table(data, colWidths=[6.5 * cm, 2 * cm, 2.5 * cm, 3.5 * cm, 3.5 * cm])

    table_style = [
        ('BACKGROUND', (0, 0), (-1, 0), HIK_DARK),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, HIK_LIGHT]),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('LEFTPADDING', (0, 1), (0, -1), 6),
        ('TOPPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
        ('LINEBELOW', (0, 0), (-1, -1), 0.5, HIK_BORDER),
    ]
    products_table.setStyle(TableStyle(table_style))
    elements.append(products_table)
    elements.append(Spacer(1, 0.5 * cm))

    total_price = order.get_total_price()
    total_price_with_vat = order.get_total_price_with_vat()
    additional_expenses_amount = order.get_additional_expenses_amount()
    total_sum = total_price_with_vat + additional_expenses_amount

    totals_data = [
        [
            Paragraph('<font name="Roboto" size="7" color="#666666">Итого без НДС</font>',
                      ParagraphStyle('', fontName='Roboto')),
            Paragraph(f'<font name="Roboto" size="7" color="#1A1A1A"><b>{total_price:.2f}</b></font>',
                      ParagraphStyle('', fontName='Roboto', alignment=TA_RIGHT)),
        ],
        [
            Paragraph('<font name="Roboto" size="7" color="#666666">Итого с НДС</font>',
                      ParagraphStyle('', fontName='Roboto')),
            Paragraph(f'<font name="Roboto" size="7" color="#1A1A1A"><b>{total_price_with_vat:.2f}</b></font>',
                      ParagraphStyle('', fontName='Roboto', alignment=TA_RIGHT)),
        ],
        [
            Paragraph('<font name="Roboto" size="7" color="#666666">Прочие расходы</font>',
                      ParagraphStyle('', fontName='Roboto')),
            Paragraph(f'<font name="Roboto" size="7" color="#1A1A1A"><b>{additional_expenses_amount:.2f}</b></font>',
                      ParagraphStyle('', fontName='Roboto', alignment=TA_RIGHT)),
        ],
        [
            Paragraph('<font name="Roboto" size="8" color="#FFFFFF"><b>Общий итог</b></font>',
                      ParagraphStyle('', fontName='Roboto')),
            Paragraph(f'<font name="Roboto" size="8" color="#FFFFFF"><b>{total_sum:.2f}</b></font>',
                      ParagraphStyle('', fontName='Roboto', alignment=TA_RIGHT)),
        ],
    ]

    totals_wrapper = [['', Table(
        totals_data,
        colWidths=[5 * cm, 4 * cm],
        style=[
            ('BACKGROUND', (0, 0), (-1, 2), colors.white),
            ('BACKGROUND', (0, 3), (-1, 3), HIK_RED),
            ('LINEBELOW', (0, 0), (-1, 2), 0.5, HIK_BORDER),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('BOX', (0, 0), (-1, -1), 0.5, HIK_BORDER),
        ]
    )]]

    totals_outer = Table(totals_wrapper, colWidths=[9 * cm, 9 * cm])
    totals_outer.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(totals_outer)
    elements.append(Spacer(1, 0.8 * cm))

    elements.append(HRFlowable(width='100%', thickness=0.5, color=HIK_BORDER, spaceAfter=5))

    footer_data = [[
        Table([
            [Paragraph('<font name="Roboto" size="5" color="#999999">ОТВЕТСТВЕННЫЙ СПЕЦИАЛИСТ</font>',
                       ParagraphStyle('', fontName='Roboto'))],
            [Paragraph('<font name="Roboto" size="9" color="#1A1A1A"><b>Хван Руслан</b></font>',
                       ParagraphStyle('', fontName='Roboto'))],
            [Paragraph('<font name="Roboto" size="7" color="#E2001A">Специалист</font>',
                       ParagraphStyle('', fontName='Roboto'))],
        ], colWidths=[9 * cm], style=[('LEFTPADDING', (0, 0), (-1, -1), 0), ('TOPPADDING', (0, 0), (-1, -1), 1),
                                      ('BOTTOMPADDING', (0, 0), (-1, -1), 1)]),
        Table([
            [Paragraph('<font name="Roboto" size="9" color="#1A1A1A"><b>RHIK</b></font>',
                       ParagraphStyle('', fontName='Roboto', alignment=TA_RIGHT))],
            [Paragraph('<font name="Roboto" size="6" color="#BBBBBB">Security &amp; Systems</font>',
                       ParagraphStyle('', fontName='Roboto', alignment=TA_RIGHT))],
        ], colWidths=[9 * cm], style=[('ALIGN', (0, 0), (-1, -1), 'RIGHT'), ('TOPPADDING', (0, 0), (-1, -1), 1),
                                      ('BOTTOMPADDING', (0, 0), (-1, -1), 1)]),
    ]]

    footer_table = Table(footer_data, colWidths=[9 * cm, 9 * cm])
    footer_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
    ]))
    elements.append(footer_table)

    doc.build(elements)
    return pdf_file_path


def send_order_to_telegram(order, file_type='excel'):
    if file_type == 'pdf':
        file_path = generate_order_pdf(order)
    else:
        file_path = generate_order_excel(order)

    url = f"https://api.telegram.org/bot{TELEGRAM_BOT_TOKEN}/sendDocument"
    with open(file_path, 'rb') as file:
        response = requests.post(url, data={'chat_id': TELEGRAM_CHAT_ID}, files={'document': file})

    os.remove(file_path)
    return response.status_code == 200
